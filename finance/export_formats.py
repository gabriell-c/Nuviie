"""Exportação formatada XLSX e PDF do financeiro."""

from __future__ import annotations

import io
from decimal import Decimal

from django.utils import timezone
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def _br_money(val) -> str:
    return f'R$ {float(val):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')


def _hex_to_rgb(hex_color: str) -> tuple:
    h = (hex_color or '#6366f1').lstrip('#')
    if len(h) == 6:
        return tuple(int(h[i:i + 2], 16) for i in (0, 2, 4))
    return 99, 102, 241


def _period_label(params) -> str:
    period = (params.get('period') or 'all').strip().lower()
    labels = {
        'all': 'Todos os períodos',
        'today': 'Hoje',
        '3d': 'Últimos 3 dias',
        '7d': 'Últimos 7 dias',
        '30d': 'Últimos 30 dias',
        '90d': 'Últimos 90 dias',
        '365d': 'Últimos 365 dias',
        'single': f"Data: {params.get('date_on', '')}",
        'range': f"{params.get('date_from', '')} → {params.get('date_to', '')}",
    }
    return labels.get(period, 'Período personalizado')


def export_styled_xlsx(entries, request_params) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Financeiro'

    thin = Side(style='thin', color='E2E8F0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:I1')
    ws['A1'] = 'NUVIIE — Controle Financeiro'
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='4F46E5')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:I2')
    ws['A2'] = f"Exportado em {timezone.localtime().strftime('%d/%m/%Y %H:%M')} · {_period_label(request_params)}"
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='64748B')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 20

    income = sum(e.amount for e in entries if e.entry_type == 'income' and e.status == 'confirmed')
    expense = sum(e.amount for e in entries if e.entry_type == 'expense' and e.status == 'confirmed')
    pending = sum(e.amount for e in entries if e.entry_type == 'income' and e.status == 'pending')

    summary_row = 4
    for col, (label, val, color) in enumerate([
        ('Entradas', income, '10B981'),
        ('Despesas', expense, 'EF4444'),
        ('Saldo', income - expense, '6366F1'),
        ('A receber', pending, 'F59E0B'),
    ], start=1):
        ws.cell(row=summary_row, column=col, value=label).font = Font(bold=True, size=9, color='64748B')
        ws.cell(row=summary_row, column=col).alignment = Alignment(horizontal='center')
        vc = ws.cell(row=summary_row + 1, column=col, value=float(val))
        vc.number_format = 'R$ #,##0.00'
        vc.font = Font(bold=True, size=14, color=color)
        vc.alignment = Alignment(horizontal='center')
        vc.fill = PatternFill('solid', fgColor='F8FAFC')

    headers = ['Data', 'Título', 'Tipo', 'Categoria', 'Valor (R$)', 'Status', 'Cliente', 'Vencimento', 'Origem']
    header_row = 7
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=ci, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = PatternFill('solid', fgColor='1E293B')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
    ws.row_dimensions[header_row].height = 22

    by_cat: dict[str, list] = {}
    for e in entries:
        by_cat.setdefault(e.category.name, []).append(e)

    row = header_row + 1
    for cat_name in sorted(by_cat.keys()):
        cat_entries = by_cat[cat_name]
        cat = cat_entries[0].category
        r, g, b = _hex_to_rgb(cat.color)
        fill_hex = f'{r:02X}{g:02X}{b:02X}'

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        cat_cell = ws.cell(row=row, column=1, value=f'  {cat_name.upper()}')
        cat_cell.font = Font(bold=True, size=11, color='FFFFFF')
        cat_cell.fill = PatternFill('solid', fgColor=fill_hex)
        cat_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 24
        row += 1

        for i, e in enumerate(cat_entries):
            tipo = 'Entrada' if e.entry_type == 'income' else 'Despesa'
            vals = [
                e.date.strftime('%d/%m/%Y'), e.title, tipo, e.category.name,
                float(e.amount), e.get_status_display(),
                e.lead.name if e.lead else '—',
                e.due_date.strftime('%d/%m/%Y') if e.due_date else '—',
                e.get_source_display(),
            ]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=val)
                c.border = border
                c.font = Font(size=10)
                if i % 2 == 0:
                    c.fill = PatternFill('solid', fgColor='F1F5F9')
                if ci == 5:
                    c.number_format = 'R$ #,##0.00'
                    c.font = Font(
                        size=10, bold=True,
                        color='10B981' if e.entry_type == 'income' else 'EF4444',
                    )
                    c.alignment = Alignment(horizontal='right')
            row += 1
        row += 1

    for i, w in enumerate([12, 32, 10, 18, 14, 12, 22, 12, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A8'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_styled_pdf(entries, request_params) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'FinTitle', parent=styles['Heading1'],
        fontSize=18, textColor=colors.HexColor('#4F46E5'), spaceAfter=4, alignment=TA_LEFT,
    )
    sub_style = ParagraphStyle(
        'FinSub', parent=styles['Normal'],
        fontSize=9, textColor=colors.HexColor('#64748B'), spaceAfter=12,
    )

    income = sum(e.amount for e in entries if e.entry_type == 'income' and e.status == 'confirmed')
    expense = sum(e.amount for e in entries if e.entry_type == 'expense' and e.status == 'confirmed')
    pending = sum(e.amount for e in entries if e.entry_type == 'income' and e.status == 'pending')

    story = [
        Paragraph('Nuviie — Controle Financeiro', title_style),
        Paragraph(
            f"Gerado em {timezone.localtime().strftime('%d/%m/%Y %H:%M')} · {_period_label(request_params)} · {len(entries)} lançamento(s)",
            sub_style,
        ),
    ]

    summary_data = [
        ['Entradas', 'Despesas', 'Saldo', 'A receber'],
        [_br_money(income), _br_money(expense), _br_money(income - expense), _br_money(pending)],
    ]
    sum_table = Table(summary_data, colWidths=[60 * mm] * 4)
    sum_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 12),
        ('TEXTCOLOR', (0, 1), (0, 1), colors.HexColor('#10B981')),
        ('TEXTCOLOR', (1, 1), (1, 1), colors.HexColor('#EF4444')),
        ('TEXTCOLOR', (2, 1), (2, 1), colors.HexColor('#6366F1')),
        ('TEXTCOLOR', (3, 1), (3, 1), colors.HexColor('#F59E0B')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, 1), [colors.HexColor('#F8FAFC')]),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story += [sum_table, Spacer(1, 10)]

    table_data = [['Data', 'Título', 'Tipo', 'Categoria', 'Valor', 'Status', 'Cliente']]
    for e in entries:
        tipo = 'Entrada' if e.entry_type == 'income' else 'Despesa'
        sign = '+' if e.entry_type == 'income' else '-'
        table_data.append([
            e.date.strftime('%d/%m/%Y'), e.title[:45], tipo, e.category.name,
            f'{sign} {_br_money(e.amount)}', e.get_status_display(),
            (e.lead.name[:28] if e.lead else '—'),
        ])
    if len(table_data) == 1:
        table_data.append(['—', 'Nenhum lançamento', '', '', '', '', ''])

    t = Table(table_data, colWidths=[22 * mm, 65 * mm, 18 * mm, 32 * mm, 28 * mm, 22 * mm, 45 * mm], repeatRows=1)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]
    for i, e in enumerate(entries, start=1):
        if i % 2 == 0:
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8FAFC')))
        color = '#10B981' if e.entry_type == 'income' else '#EF4444'
        style_cmds.append(('TEXTCOLOR', (4, i), (4, i), colors.HexColor(color)))
    t.setStyle(TableStyle(style_cmds))
    story.append(t)
    doc.build(story)
    return buf.getvalue()
